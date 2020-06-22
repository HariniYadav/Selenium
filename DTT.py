import openpyxl
path="C:\\Users\\Narayana\Desktop\Harini\Book1.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active
#sheet=workbook.get_sheet_by_name("Sheet1")

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="           ")
    print()

sheet1=workbook["Sheet2"]
for r in range(1,6):
    for c in range(1,4):
        sheet1.cell(r,c).value="welcome"

workbook.save(path)